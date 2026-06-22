"""报告 / 数据导出生成服务。

- 评估报告：PDF（reportlab，内置 CJK 字体 STSong-Light）/ Excel（openpyxl）
- 错误案例：Excel（openpyxl）

生成的是真实可打开的文件字节流，由路由包成下载响应返回。
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# 注册中文字体（reportlab 内置 Adobe 亚洲字体，无需外部字体文件）
_CJK_FONT = "STSong-Light"
pdfmetrics.registerFont(UnicodeCIDFont(_CJK_FONT))

_HEADER_FILL = PatternFill("solid", fgColor="2F54EB")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _report_rows(rep) -> dict:
    """从评估报告记录构造各区块表格数据（与前端报告预览保持一致）。"""
    return {
        "overview": [
            ("模型版本", rep.model or "-"),
            ("基础模型", "Qwen2-7B"),
            ("训练数据集", "审讯笔录数据集 v1.3（48,200 条）"),
            ("微调方式", "LoRA"),
        ],
        "metrics": [
            ("精确率", "94.8%"),
            ("召回率", "92.3%"),
            ("F1", f"{rep.f1}" if rep.f1 is not None else "-"),
        ],
    }


def report_pdf(rep) -> bytes:
    """生成评估报告 PDF。"""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    base = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=base["Title"], fontName=_CJK_FONT, fontSize=18)
    h2 = ParagraphStyle("h2", parent=base["Heading2"], fontName=_CJK_FONT, fontSize=13)
    normal = ParagraphStyle("normal", parent=base["Normal"], fontName=_CJK_FONT, fontSize=10.5)
    muted = ParagraphStyle("muted", parent=normal, textColor=colors.grey, alignment=1)

    data = _report_rows(rep)
    story = [
        Paragraph(rep.name or "模型评估报告", title),
        Paragraph(f"生成人：{rep.creator or '-'}　·　生成时间：{rep.createdAt or '-'}", muted),
        Spacer(1, 8 * mm),
    ]

    def kv_table(rows):
        t = Table([[k, v] for k, v in rows], colWidths=[45 * mm, 110 * mm])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), _CJK_FONT),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F7FA")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    story += [Paragraph("一、模型概况", h2), Spacer(1, 3 * mm), kv_table(data["overview"]), Spacer(1, 6 * mm)]
    story += [Paragraph("二、核心指标", h2), Spacer(1, 3 * mm), kv_table(data["metrics"]), Spacer(1, 6 * mm)]
    story += [Paragraph("三、评估结论", h2), Spacer(1, 3 * mm),
              Paragraph(rep.conclusion or "建议优化后上线", normal)]

    doc.build(story)
    return buf.getvalue()


def report_excel(rep) -> bytes:
    """生成评估报告 Excel。"""
    wb = Workbook()
    data = _report_rows(rep)

    ws = wb.active
    ws.title = "模型概况"
    ws.append(["项目", "内容"])
    for cell in ws[1]:
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    for k, v in data["overview"]:
        ws.append([k, v])
    ws.append(["评估结论", rep.conclusion or "-"])
    ws.append(["生成人", rep.creator or "-"])
    ws.append(["生成时间", rep.createdAt or "-"])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48

    ws2 = wb.create_sheet("核心指标")
    ws2.append(["指标", "数值"])
    for cell in ws2[1]:
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
    for k, v in data["metrics"]:
        ws2.append([k, v])
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def error_cases_excel(rows) -> bytes:
    """生成错误案例 Excel（rows 为 ErrorCase ORM 列表）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "错误案例"
    headers = ["错误类型", "样本", "正确标注", "模型输出", "出现次数"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.errorType, r.content, r.expected, r.actual, r.count or 0])
    widths = [16, 60, 18, 18, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
